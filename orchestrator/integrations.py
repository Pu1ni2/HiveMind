"""
Real-world integrations for HIVEMIND agents.

These functions let forged tools ACTUALLY interact with external services:
send emails, post to Slack, create calendar events, parse resumes,
generate spreadsheets, trigger webhooks, and read/create PDFs.

All functions are designed to:
  1. Work gracefully when credentials aren't configured (return helpful error)
  2. Be safe to call from LLM-generated code
  3. Return strings (consumed by agents as text)
"""

import html as html_module
import os
import re
import json
import csv
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta

from .capabilities import save_file, OUTPUT_DIR

# =====================================================================
# CONFIG — loaded from env vars
# =====================================================================

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "") or SMTP_USER

SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL", "")

GOOGLE_CALENDAR_CREDENTIALS = os.getenv("GOOGLE_CALENDAR_CREDENTIALS", "")


# =====================================================================
# EMAIL — send real emails via SMTP
# =====================================================================

def send_email(
    to: str,
    subject: str,
    body: str,
    cc: str = "",
    html: bool = False,
) -> str:
    """Send a REAL email via SMTP.

    Parameters:
        to: recipient email (comma-separated for multiple)
        subject: email subject line
        body: email body (plain text or HTML)
        cc: CC recipients (comma-separated)
        html: if True, body is treated as HTML

    Returns: success/error message string.
    """
    if not SMTP_USER or not SMTP_PASS:
        # Save as draft instead
        draft = f"To: {to}\nCC: {cc}\nSubject: {subject}\n\n{body}"
        filename = f"email_draft_{_safe_filename(subject)}.txt"
        save_file(filename, draft)
        return (
            f"Email credentials not configured (set SMTP_USER, SMTP_PASS in .env). "
            f"Saved email draft to output/{filename}. "
            f"To enable real email sending, add SMTP_USER and SMTP_PASS to your .env file."
        )

    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = SMTP_FROM
        msg["To"] = to
        msg["Subject"] = subject
        if cc:
            msg["Cc"] = cc

        if html:
            msg.attach(MIMEText(body, "html"))
        else:
            msg.attach(MIMEText(body, "plain"))

        all_recipients = [addr.strip() for addr in to.split(",")]
        if cc:
            all_recipients += [addr.strip() for addr in cc.split(",")]

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_FROM, all_recipients, msg.as_string())

        return f"Email sent successfully to {to}" + (f" (CC: {cc})" if cc else "")

    except Exception as e:
        return f"Email send error: {e}"


# =====================================================================
# SLACK — post messages to Slack via webhook
# =====================================================================

def send_slack_message(
    message: str,
    channel: str = "",
    blocks: list | None = None,
) -> str:
    """Send a REAL message to Slack via incoming webhook.

    Parameters:
        message: the message text (supports Slack markdown)
        channel: optional channel override
        blocks: optional Slack Block Kit blocks (list of dicts)

    Returns: success/error message string.

    Setup: Create a Slack incoming webhook at
    https://api.slack.com/messaging/webhooks
    and set SLACK_WEBHOOK_URL in .env
    """
    import requests

    if not SLACK_WEBHOOK_URL:
        filename = f"slack_message_{datetime.now().strftime('%H%M%S')}.md"
        save_file(filename, f"**Slack Message**\n\n{message}")
        return (
            f"Slack webhook not configured (set SLACK_WEBHOOK_URL in .env). "
            f"Saved message to output/{filename}. "
            f"To enable real Slack messaging, create an incoming webhook and add SLACK_WEBHOOK_URL to .env."
        )

    payload = {"text": message}
    if channel:
        payload["channel"] = channel
    if blocks:
        payload["blocks"] = blocks

    try:
        resp = requests.post(
            SLACK_WEBHOOK_URL,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=10,
        )
        if resp.status_code == 200:
            return f"Slack message sent successfully"
        return f"Slack error (HTTP {resp.status_code}): {resp.text}"
    except Exception as e:
        return f"Slack send error: {e}"


# =====================================================================
# CALENDAR — create .ics calendar events (universal format)
# =====================================================================

def create_calendar_event(
    title: str,
    start: str,
    end: str = "",
    description: str = "",
    location: str = "",
    attendees: str = "",
) -> str:
    """Create a REAL .ics calendar file that can be imported into
    Google Calendar, Outlook, Apple Calendar, or any calendar app.

    Parameters:
        title: event title
        start: start datetime (ISO format or "YYYY-MM-DD HH:MM")
        end: end datetime (if empty, defaults to start + 1 hour)
        description: event description
        location: event location
        attendees: comma-separated email addresses

    Returns: path to the saved .ics file.
    """
    try:
        # Parse start time
        start_dt = _parse_datetime(start)
        if not start_dt:
            return f"Invalid start time: {start}. Use format: YYYY-MM-DD HH:MM"

        # Parse or default end time
        if end:
            end_dt = _parse_datetime(end)
            if not end_dt:
                return f"Invalid end time: {end}"
        else:
            end_dt = start_dt + timedelta(hours=1)

        # Build .ics content
        uid = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{_safe_filename(title)}@hivemind"
        now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

        ics_lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//HIVEMIND//Agent Orchestration//EN",
            "CALSCALE:GREGORIAN",
            "METHOD:PUBLISH",
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}",
            f"DTSTAMP:{now}",
            f"SUMMARY:{_ics_escape(title)}",
        ]

        if description:
            ics_lines.append(f"DESCRIPTION:{_ics_escape(description)}")
        if location:
            ics_lines.append(f"LOCATION:{_ics_escape(location)}")

        if attendees:
            for email in attendees.split(","):
                email = email.strip()
                if email:
                    ics_lines.append(f"ATTENDEE;RSVP=TRUE:mailto:{email}")

        ics_lines += [
            "STATUS:CONFIRMED",
            "END:VEVENT",
            "END:VCALENDAR",
        ]

        ics_content = "\r\n".join(ics_lines)
        filename = f"{_safe_filename(title)}.ics"
        result = save_file(filename, ics_content)

        return (
            f"Calendar event created: {filename}\n"
            f"  Title: {title}\n"
            f"  Start: {start_dt}\n"
            f"  End: {end_dt}\n"
            f"  Location: {location or 'N/A'}\n"
            f"  Attendees: {attendees or 'None'}\n"
            f"  File: {result}\n\n"
            f"Import this .ics file into Google Calendar, Outlook, or Apple Calendar."
        )

    except Exception as e:
        return f"Calendar event creation error: {e}"


# =====================================================================
# RESUME / DOCUMENT ANALYSIS
# =====================================================================

def parse_resume(text: str) -> str:
    """Analyze resume/CV text and extract structured information.

    Parameters:
        text: the full text of a resume

    Returns: structured analysis as formatted string.
    """
    analysis = {
        "emails": re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', text),
        "phones": re.findall(r'[\+]?[\d\s\-\(\)]{10,}', text),
        "urls": re.findall(r'https?://\S+|www\.\S+|linkedin\.com/\S+|github\.com/\S+', text, re.I),
    }

    # Extract sections
    section_patterns = {
        "education": r'(?i)(education|academic|university|degree|bachelor|master|phd|school)[\s\S]{0,2000}?(?=\n\s*\n[A-Z]|\Z)',
        "experience": r'(?i)(experience|employment|work\s*history|professional)[\s\S]{0,3000}?(?=\n\s*\n[A-Z]|\Z)',
        "skills": r'(?i)(skills|technologies|tech\s*stack|proficien|competenc)[\s\S]{0,1500}?(?=\n\s*\n[A-Z]|\Z)',
        "certifications": r'(?i)(certific|license|accredit)[\s\S]{0,1000}?(?=\n\s*\n[A-Z]|\Z)',
    }

    for section_name, pattern in section_patterns.items():
        match = re.search(pattern, text)
        analysis[section_name] = match.group(0).strip()[:800] if match else "Not found"

    # Count years of experience (rough estimate)
    years = re.findall(r'(\d{4})\s*[-–]\s*(\d{4}|present|current)', text, re.I)
    total_years = 0
    for start_y, end_y in years:
        try:
            start = int(start_y)
            end = datetime.now().year if end_y.lower() in ('present', 'current') else int(end_y)
            total_years += max(0, end - start)
        except ValueError:
            pass

    result = "# Resume Analysis\n\n"
    result += f"## Contact Information\n"
    result += f"- Email(s): {', '.join(analysis['emails']) or 'Not found'}\n"
    result += f"- Phone(s): {', '.join(analysis['phones']) or 'Not found'}\n"
    result += f"- Links: {', '.join(analysis['urls']) or 'Not found'}\n"
    result += f"- Estimated Experience: ~{total_years} years\n\n"

    for section in ['education', 'experience', 'skills', 'certifications']:
        result += f"## {section.title()}\n{analysis[section]}\n\n"

    return result


def read_pdf(filepath: str) -> str:
    """Read text content from a PDF file.

    Parameters:
        filepath: path to the PDF file (can be in output/ dir or absolute)

    Returns: extracted text from the PDF.
    """
    # Resolve path
    if not os.path.isabs(filepath):
        filepath = os.path.join(OUTPUT_DIR, filepath)

    if not os.path.exists(filepath):
        return f"PDF file not found: {filepath}"

    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        if not text_parts:
            return "PDF exists but no text could be extracted (might be image-based)."
        return "\n\n---\n\n".join(text_parts)
    except ImportError:
        # Fallback: try PyPDF2
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            text_parts = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            if not text_parts:
                return "PDF exists but no text could be extracted."
            return "\n\n---\n\n".join(text_parts)
        except ImportError:
            return (
                "PDF reading requires pdfplumber or PyPDF2. "
                "Install with: pip install pdfplumber"
            )
    except Exception as e:
        return f"PDF read error: {e}"


# =====================================================================
# SPREADSHEETS — create real CSV and Excel files
# =====================================================================

def create_spreadsheet(
    filename: str,
    headers: list,
    rows: list,
    sheet_name: str = "Sheet1",
) -> str:
    """Create a REAL spreadsheet file (CSV or Excel).

    Parameters:
        filename: output filename (use .csv or .xlsx extension)
        headers: list of column header strings
        rows: list of lists (each inner list is a row of values)
        sheet_name: sheet name for Excel files

    Returns: path to the saved file.
    """
    if filename.endswith(".xlsx"):
        return _create_excel(filename, headers, rows, sheet_name)
    else:
        return _create_csv(filename, headers, rows)


def _create_csv(filename: str, headers: list, rows: list) -> str:
    if not filename.endswith(".csv"):
        filename += ".csv"
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        content = output.getvalue()
        return save_file(filename, content)
    except Exception as e:
        return f"CSV creation error: {e}"


def _create_excel(filename: str, headers: list, rows: list, sheet_name: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 50)

        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        return f"Excel file saved: {filepath} ({len(rows)} rows)"

    except ImportError:
        # Fallback to CSV
        csv_name = filename.replace(".xlsx", ".csv")
        result = _create_csv(csv_name, headers, rows)
        return f"openpyxl not installed — saved as CSV instead. {result}"
    except Exception as e:
        return f"Excel creation error: {e}"


# =====================================================================
# WEBHOOKS — trigger any external service
# =====================================================================

def send_webhook(
    url: str,
    payload: dict,
    method: str = "POST",
    headers: dict | None = None,
) -> str:
    """Send a REAL HTTP webhook to any external service.

    Parameters:
        url: the webhook URL
        payload: JSON payload dict
        method: HTTP method (POST, PUT, PATCH)
        headers: optional extra headers

    Returns: response status and body.
    """
    import requests

    if not url.startswith("http"):
        return "Invalid webhook URL — must start with http:// or https://"

    req_headers = {"Content-Type": "application/json"}
    if headers:
        req_headers.update(headers)

    try:
        resp = requests.request(
            method=method.upper(),
            url=url,
            json=payload,
            headers=req_headers,
            timeout=15,
        )
        return (
            f"Webhook response ({resp.status_code}):\n"
            f"{resp.text[:2000]}"
        )
    except Exception as e:
        return f"Webhook error: {e}"


# =====================================================================
# TRELLO / KANBAN — create project boards
# =====================================================================

def create_kanban_board(
    title: str,
    columns: list[dict],
) -> str:
    """Create a REAL interactive Kanban board as an HTML file.

    Parameters:
        title: board title
        columns: list of {"name": "To Do", "cards": [{"title": "...", "desc": "...", "tag": "..."}]}

    Returns: path to the saved HTML file.
    """
    # Escape all user-controlled strings before embedding in HTML to prevent XSS.
    safe_title = html_module.escape(title)

    cols_html = ""
    for col in columns:
        cards_html = ""
        for card in col.get("cards", []):
            card_title = html_module.escape(str(card.get("title", "")))
            card_desc = html_module.escape(str(card.get("desc", "")))
            card_tag = html_module.escape(str(card.get("tag", "")))
            tag_html = f'<span class="card-tag">{card_tag}</span>' if card_tag else ""
            cards_html += f"""
            <div class="card" draggable="true">
                <div class="card-title">{card_title}</div>
                <div class="card-desc">{card_desc}</div>
                {tag_html}
            </div>"""

        col_name = html_module.escape(str(col.get("name", "Column")))
        cols_html += f"""
        <div class="column" ondragover="event.preventDefault()" ondrop="drop(event)">
            <div class="col-header">{col_name} <span class="col-count">{len(col.get('cards', []))}</span></div>
            <div class="cards">{cards_html}</div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{safe_title} — HIVEMIND Board</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0a0a0b;color:#e4e4e7;min-height:100vh;padding:32px}}
h1{{font-size:1.5rem;font-weight:700;margin-bottom:24px;background:linear-gradient(135deg,#818cf8,#a78bfa);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.board{{display:flex;gap:16px;overflow-x:auto;padding-bottom:20px}}
.column{{min-width:280px;max-width:320px;background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.06);border-radius:12px;padding:16px;flex-shrink:0}}
.col-header{{font-size:.85rem;font-weight:700;color:#a1a1aa;text-transform:uppercase;letter-spacing:.06em;margin-bottom:14px;display:flex;align-items:center;justify-content:space-between}}
.col-count{{background:rgba(99,102,241,.12);color:#818cf8;font-size:.7rem;padding:2px 8px;border-radius:99px}}
.cards{{display:flex;flex-direction:column;gap:10px;min-height:60px}}
.card{{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.08);border-radius:8px;padding:14px;cursor:grab;transition:transform .15s,box-shadow .15s}}
.card:hover{{transform:translateY(-2px);box-shadow:0 4px 16px rgba(0,0,0,.3)}}
.card.dragging{{opacity:.5}}
.card-title{{font-size:.88rem;font-weight:600;color:#fafafa;margin-bottom:6px}}
.card-desc{{font-size:.78rem;color:#71717a;line-height:1.5}}
.card-tag{{display:inline-block;margin-top:8px;padding:2px 8px;font-size:.68rem;font-weight:600;background:rgba(99,102,241,.1);border:1px solid rgba(99,102,241,.15);border-radius:99px;color:#818cf8}}
</style></head><body>
<h1>{safe_title}</h1>
<div class="board">{cols_html}</div>
<script>
let dragged=null;
document.querySelectorAll('.card').forEach(c=>{{
c.addEventListener('dragstart',e=>{{dragged=c;c.classList.add('dragging')}});
c.addEventListener('dragend',()=>{{dragged.classList.remove('dragging');dragged=null}});
}});
function drop(e){{e.preventDefault();if(dragged)e.currentTarget.querySelector('.cards').appendChild(dragged)}}
</script></body></html>"""

    filename = f"{_safe_filename(title)}_board.html"
    return save_file(filename, html)


# =====================================================================
# HELPERS
# =====================================================================

def _safe_filename(text: str) -> str:
    return re.sub(r'[<>:"/\\|?*\s]+', '_', text)[:60].strip('_').lower()


def _parse_datetime(s: str) -> datetime | None:
    formats = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%d-%m-%Y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return None


def _ics_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;").replace("\n", "\\n")


# =====================================================================
# INTEGRATION NAMESPACE — injected into forged tools
# =====================================================================

INTEGRATION_NAMESPACE = {
    "_send_email": send_email,
    "_send_slack": send_slack_message,
    "_create_calendar_event": create_calendar_event,
    "_parse_resume": parse_resume,
    "_read_pdf": read_pdf,
    "_create_spreadsheet": create_spreadsheet,
    "_send_webhook": send_webhook,
    "_create_kanban_board": create_kanban_board,
}
