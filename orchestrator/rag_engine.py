"""
RAG Engine — turns any agent into an interactive document Q&A system.

When the DA creates an agent with type "rag", this engine:
1. Accepts file uploads (PDF, Excel, CSV, TXT, MD)
2. Chunks and indexes the content in ChromaDB
3. Answers queries using semantic retrieval + LLM synthesis
"""

from __future__ import annotations
import hashlib
import os
import re
import json
import csv
import io
import threading

from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage

from .config import OPENAI_API_KEY
from .capabilities import OUTPUT_DIR

# Thread-safe per-agent RAG collections
_agent_collections: dict[str, object] = {}
_collections_lock = threading.Lock()
_chroma_client = None
_chroma_lock = threading.Lock()

# Relevance threshold: cosine distance > this value means the chunk is
# too dissimilar to the query and should be excluded from context.
_RELEVANCE_THRESHOLD = 0.7  # cosine distance (0=identical, 2=opposite)


def _get_chroma():
    global _chroma_client
    with _chroma_lock:
        if _chroma_client is not None:
            return _chroma_client
        try:
            import chromadb
            os.makedirs("data", exist_ok=True)
            _chroma_client = chromadb.PersistentClient(path="data/hivemind_rag")
            return _chroma_client
        except ImportError:
            return None
        except Exception:
            return None


def _get_collection(agent_id: str):
    """Get or create a ChromaDB collection for an agent (thread-safe)."""
    with _collections_lock:
        if agent_id in _agent_collections:
            return _agent_collections[agent_id]
        client = _get_chroma()
        if client is None:
            return None
        col_name = f"rag_{agent_id.replace('-', '_')[:50]}"
        col = client.get_or_create_collection(name=col_name, metadata={"hnsw:space": "cosine"})
        _agent_collections[agent_id] = col
        return col


def _content_hash(content_bytes: bytes) -> str:
    """SHA-256 hex digest of file content, used for deduplication."""
    return hashlib.sha256(content_bytes).hexdigest()[:16]


# ══════════════════════════════════════════════════════════════════
# FILE PROCESSING
# ══════════════════════════════════════════════════════════════════

def process_upload(agent_id: str, filename: str, content_bytes: bytes) -> dict:
    """Process an uploaded file: extract text, chunk, and index.

    Returns: {"status": "ok"|"error", "chunks": int, "filename": str, "message": str}
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    try:
        if ext == "pdf":
            text = _extract_pdf(content_bytes)
        elif ext in ("xlsx", "xls"):
            text = _extract_excel(content_bytes)
        elif ext == "csv":
            text = _extract_csv(content_bytes)
        elif ext in ("txt", "md", "text", "markdown"):
            text = content_bytes.decode("utf-8", errors="replace")
        elif ext == "json":
            data = json.loads(content_bytes.decode("utf-8", errors="replace"))
            text = json.dumps(data, indent=2)
        else:
            # Try as plain text
            text = content_bytes.decode("utf-8", errors="replace")

        if not text or len(text.strip()) < 10:
            return {"status": "error", "chunks": 0, "filename": filename,
                    "message": "No text could be extracted from this file."}

        # Chunk the text
        chunks = _chunk_text(text, chunk_size=800, overlap=100)

        # Index in ChromaDB
        col = _get_collection(agent_id)
        if col is None:
            return {"status": "error", "chunks": 0, "filename": filename,
                    "message": "ChromaDB not available. Install with: pip install chromadb"}

        # Use content hash in IDs so re-uploading the same file overwrites
        # existing chunks rather than creating duplicates.
        file_hash = _content_hash(content_bytes)
        ids = [f"{file_hash}_{i}" for i in range(len(chunks))]
        metadatas = [{"filename": filename, "chunk_index": i, "total_chunks": len(chunks), "file_hash": file_hash} for i in range(len(chunks))]

        # Upsert in batches (ChromaDB has batch limits)
        batch_size = 40
        for start in range(0, len(chunks), batch_size):
            end = min(start + batch_size, len(chunks))
            col.upsert(
                documents=chunks[start:end],
                ids=ids[start:end],
                metadatas=metadatas[start:end],
            )

        # Also save the raw file to output/
        save_path = os.path.join(OUTPUT_DIR, filename)
        os.makedirs(os.path.dirname(save_path) or OUTPUT_DIR, exist_ok=True)
        with open(save_path, "wb") as f:
            f.write(content_bytes)

        return {
            "status": "ok",
            "chunks": len(chunks),
            "filename": filename,
            "chars": len(text),
            "message": f"Indexed {len(chunks)} chunks from {filename} ({len(text)} chars)",
        }

    except Exception as exc:
        return {"status": "error", "chunks": 0, "filename": filename,
                "message": f"Processing error: {exc}"}


def _extract_pdf(content_bytes: bytes) -> str:
    """Extract text from PDF bytes."""
    try:
        import pdfplumber
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name
        try:
            parts = []
            with pdfplumber.open(tmp_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        parts.append(page_text)
            return "\n\n".join(parts)
        finally:
            os.unlink(tmp_path)
    except ImportError:
        return content_bytes.decode("utf-8", errors="replace")


def _extract_excel(content_bytes: bytes) -> str:
    """Extract text from Excel bytes."""
    try:
        from openpyxl import load_workbook
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name
        try:
            wb = load_workbook(tmp_path, read_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"## Sheet: {sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        parts.append(row_text)
            return "\n".join(parts)
        finally:
            os.unlink(tmp_path)
    except ImportError:
        return "Excel support requires openpyxl"


def _extract_csv(content_bytes: bytes) -> str:
    """Extract text from CSV bytes."""
    text = content_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append(" | ".join(row))
    return "\n".join(rows)


def _chunk_text(text: str, chunk_size: int = 800, overlap: int = 100) -> list[str]:
    """Split text into chunks with proper context-preserving overlap.

    Strategy:
    1. Split by paragraph boundaries first to preserve natural structure.
    2. If a paragraph exceeds chunk_size, split further by sentence boundaries.
    3. Overlap is implemented by seeding each new chunk with the tail of the
       previous chunk — the tail text is NOT duplicated within the same chunk,
       only carried forward as context seed for the next one.
    """
    overlap = max(0, min(overlap, chunk_size // 2))  # guard: overlap < half chunk

    # Split by paragraphs first
    paragraphs = re.split(r'\n\s*\n', text)
    raw_chunks: list[str] = []
    current_chunk = ""

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue

        if len(current_chunk) + len(para) + 2 <= chunk_size:
            current_chunk += ("\n\n" + para if current_chunk else para)
        else:
            if current_chunk:
                raw_chunks.append(current_chunk)
            # If the paragraph itself is too long, split by sentences
            if len(para) > chunk_size:
                sentences = re.split(r'(?<=[.!?])\s+', para)
                current_chunk = ""
                for sent in sentences:
                    sent = sent.strip()
                    if not sent:
                        continue
                    if len(current_chunk) + len(sent) + 1 <= chunk_size:
                        current_chunk += (" " + sent if current_chunk else sent)
                    else:
                        if current_chunk:
                            raw_chunks.append(current_chunk)
                        # Sentence longer than chunk_size: hard split
                        if len(sent) > chunk_size:
                            for start in range(0, len(sent), chunk_size):
                                raw_chunks.append(sent[start:start + chunk_size])
                            current_chunk = ""
                        else:
                            current_chunk = sent
            else:
                current_chunk = para

    if current_chunk:
        raw_chunks.append(current_chunk)

    if not raw_chunks:
        return [text[:chunk_size]] if text.strip() else []

    # Add overlap: each chunk[i] is seeded with the tail of chunk[i-1].
    # This gives the retriever context continuity without duplicating text
    # inside any single chunk.
    if overlap > 0 and len(raw_chunks) > 1:
        overlapped = [raw_chunks[0]]
        for i in range(1, len(raw_chunks)):
            tail = raw_chunks[i - 1][-overlap:]
            overlapped.append(tail + " " + raw_chunks[i])
        return overlapped

    return raw_chunks


# ══════════════════════════════════════════════════════════════════
# QUERY
# ══════════════════════════════════════════════════════════════════

def query_rag(
    agent_id: str,
    question: str,
    agent_role: str = "",
    agent_persona: str = "",
    agent_objective: str = "",
    n_results: int = 5,
) -> dict:
    """Query the RAG agent with a question.

    Returns: {"answer": str, "sources": list[dict], "status": str}
    """
    col = _get_collection(agent_id)
    if col is None:
        return {"answer": "RAG not available. ChromaDB not initialized.", "sources": [], "status": "error"}

    # Check if collection has any documents
    if col.count() == 0:
        return {"answer": "No documents uploaded yet. Please upload files first.", "sources": [], "status": "empty"}

    # Retrieve relevant chunks
    try:
        results = col.query(query_texts=[question], n_results=min(n_results, col.count()))
    except Exception as exc:
        return {"answer": f"Search error: {exc}", "sources": [], "status": "error"}

    chunks = results["documents"][0] if results["documents"] else []
    metadatas = results["metadatas"][0] if results["metadatas"] else []
    distances = results["distances"][0] if results["distances"] else []

    if not chunks:
        return {"answer": "No relevant content found in the uploaded documents.", "sources": [], "status": "no_results"}

    # Filter out chunks that are too dissimilar to the query.
    # ChromaDB uses cosine distance (0=identical, 2=opposite); threshold 0.7
    # excludes chunks with < 30% similarity — avoids hallucination from noise.
    filtered = [
        (chunk, meta, dist)
        for chunk, meta, dist in zip(chunks, metadatas, distances)
        if dist <= _RELEVANCE_THRESHOLD
    ]
    if not filtered:
        # All chunks exceeded threshold — fall back to top result rather than
        # returning empty (some answer is better than none for RAG queries).
        filtered = [(chunks[0], metadatas[0], distances[0])]

    # Build context
    context_parts = []
    sources = []
    for i, (chunk, meta, dist) in enumerate(filtered):
        context_parts.append(f"[Source {i+1}: {meta.get('filename', '?')}, chunk {meta.get('chunk_index', '?')}]\n{chunk}")
        sources.append({
            "filename": meta.get("filename", "?"),
            "chunk_index": meta.get("chunk_index", 0),
            "relevance": round(1 - dist, 3),
            "preview": chunk[:150],
        })

    context = "\n\n---\n\n".join(context_parts)

    # Build system prompt with agent identity
    system = f"You are {agent_role}.\n\n{agent_persona}\n\nObjective: {agent_objective}\n\n"
    system += (
        "Answer the user's question based ONLY on the retrieved document context below. "
        "If the answer is not in the context, say so. "
        "Cite sources by referring to [Source N]. "
        "Be thorough, specific, and actionable.\n\n"
        f"RETRIEVED CONTEXT:\n{context}"
    )

    # Query LLM
    model = ChatOpenAI(model="gpt-4o", api_key=OPENAI_API_KEY, temperature=0.3)
    try:
        response = model.invoke([
            SystemMessage(content=system),
            HumanMessage(content=question),
        ])
        answer = response.content
    except Exception as exc:
        answer = f"LLM error: {exc}"

    return {"answer": answer, "sources": sources, "status": "ok"}


def get_agent_files(agent_id: str) -> list[dict]:
    """List all files indexed for an agent."""
    col = _get_collection(agent_id)
    if col is None or col.count() == 0:
        return []

    # Get unique filenames from metadata
    try:
        all_data = col.get(include=["metadatas"])
        files = {}
        for meta in all_data["metadatas"]:
            fname = meta.get("filename", "?")
            if fname not in files:
                files[fname] = {"filename": fname, "chunks": 0}
            files[fname]["chunks"] += 1
        return list(files.values())
    except Exception:
        return []
